#! python3
# A Shebang indicates which Bash or Python interpreter to use to interpret an executable file.
# change formats in an XLSX worksheet 
# https://www.geeksforgeeks.org/formatting-cells-using-openpyxl-in-python/

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

outputPath = "C:/Users/ben/OneDrive/Documents/Gaba_Docs/Whisperer/FBRef/2025 Files/AllDomesticLeagues/"
outputFile = "Formats_Chg" + ".xlsx"

# Create a workbook
workbook = Workbook()

# Get the active sheet
sheet = workbook.active

# Change content in cell
sheet['A1'] = "hello"
sheet['B1'] = "geek"

sheet.title = "Geek Sheet"

# Format number to 2 decimal places
sheet['A3'] = 4312345.6789
sheet['A3'].number_format = '#,##0.00'

# Format date
sheet['B3'] = "2024-01-01"
sheet['B3'].number_format = 'yyyy-mm-dd'

# Example: Tomato-colored, bold, and italic text
sheet['A1'].font = Font(
  					name='Calibri',
  					bold=True,
  					italic=True,
  					size=14,
  					color="FF6347"
				)


# Example: Solid yellow fill
sheet['B1'].fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

# Change the allignment
sheet['C1'] = "I'm vertically and Horizontally Centered"
sheet['C1'].alignment = Alignment(horizontal='center', vertical='center')

# Set border properties
border_thick = Side(style='thick')
border_thin = Side(style='thin')
borded_dashed = Side(style='dashed')
border_dotted = Side(style='dotted')

# Apply different border styles
sheet['C1'].border = Border(top=border_thick, left=border_thin, right=borded_dashed, bottom=border_dotted)

# for col in range(3, ws.max_column+1):
#    for row in range(1, last_cell):
#        ws.cell(column=col, row=row).number_format = '@'  # Changing format to TEXT

# Save the work book
workbook.save(outputPath+outputFile)