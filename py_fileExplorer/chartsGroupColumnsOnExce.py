#! python3
# A Shebang indicates which Bash or Python interpreter to use to interpret an executable file. 
# https://app.datacamp.com/learn/tutorials/python-excel-tutorial
import sys, webbrowser, re, time , os
from pathlib import Path
# add a day to a date field
from datetime import datetime, timedelta, date
import openpyxl
from openpyxl.chart import Reference  # Values for plotting
from openpyxl.chart import BarChart


inputPath = "C:/Users/ben/OneDrive/Documents/Gaba_Docs/Whisperer/FBRef/2025 Files/AllDomesticLeagues/"
inputFile = "Fbref_20250102"   + ".xlsx"

outputPath = "C:/Users/ben/OneDrive/Documents/Gaba_Docs/development/Python/Python_Ben/pgmOutput/"
outputFile = "WSB_20250102a" + ".xlsx"

iterate =[]
iterate.append('one')

for ite in iterate:

    try:
        # Load workbook
        wb = openpyxl.load_workbook(inputPath+inputFile)

        # Select active sheet
        # ws = wb.active

        # Select specific sheet 
        ws = wb['Fbref_Charts_GrpCols']

        # count the number of rows and columns in this worksheet
        print('Total number of rows: '+str(ws.max_row)+'. And total number of columns: '+str(ws.max_column))

        # Access cell values / retrieve data from a specific cell 
        cell_value = ws['B2'].value

        # barChart

        # Categories
        # 'Fbref_Charts'!$A$2:$A$26
        values = Reference(ws,         # worksheet object   
                        min_col=2,  # minimum column where your values begin
                        max_col=5,  # maximum column where your values end
                        min_row=1,  # minimum row you’d like to plot from
                        max_row=26) # maximum row you’d like to plot from
        
        cats = Reference(ws,         # worksheet object   
                min_col=1,  # minimum column where your values begin
                max_col=1,  # maximum column where your values end
                min_row=2,  # minimum row you’d like to plot from
                max_row=26) # maximum row you’d like to plot from
        
        # CreaTE Bar chart
        chart = BarChart()
        chart.add_data(values, titles_from_data=True)
        chart.set_categories(cats)

        # Set Chart Titles
        # set the title of the chart
        chart.title = "Total Match Played (MP)"

        # set the title of the x-axis
        chart.x_axis.title = "Competion/League"

        # set the title of the y-axis
        chart.y_axis.title = "Total MP Break down by W/L/D"

        chart.style="10.0"

        # the top-left corner of the chart
        # is anchored to cell F2 .
        ws.add_chart(chart,"J2")

        #  
        print("Bar Chart X Values >>>     " + '\n' + str(cats))
        print("Bar Chart Y Values >>>     " + '\n' + str(values))
        #(print ("{:>10}{:>55}".format("Before >>>" ,  str(wb.sheetnames))))

        # answer = [i for i in wb.sheetnames if 'Copy' in i] #return if part of a string is contain

        break


    except Exception as e:
        print("could not open input file..: " + inputPath+inputFile+'\n' + inputPath + '\n' + inputFile + '\n' +str(e))

wb.save(inputPath+inputFile)
wb.close()
exit