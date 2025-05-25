#! python3
# A Shebang indicates which Bash or Python interpreter to use to interpret an executable file. 

# https://www.geeksforgeeks.org/check-a-file-is-opened-or-closed-in-python/
# https://www.youtube.com/watch?v=Uq8upBnfzPg

import openpyxl

inputPath = "C:/Users/ben/OneDrive/Documents/Gaba_Docs/Whisperer/FBRef/2025 Files/AllDomesticLeagues/"
inputFile = "Fbref_Teams"   + ".xlsx"

# function to check whether the file is closed
def checkFileClosed(file_obj):
    # check if file is closed using `closed` property
    if file_obj.closed == True:
        print('Your file is closed.')
    else:
        print('Your file is open.')
    
    return file_obj.closed

def checkFileClosedOpnxlpy(file_obj2):
    if file_obj2.active == "FBref_Teams":
        return True
        # print('Your file is closed.')
    else:
        return False
        # print('Your file is copen.')

    # check if file is closed using `closed` property
    if file_obj2.sheetnames == []:
        return True
        # print('Your file is closed.')
    else:
        return False
        # print('Your file is copen.')

    if file_obj2.active == []:
        return True
        # print('Your file is closed.')
    else:
        return False
        # print('Your file is copen.')
    

# Opening the "Fbref_Teams" file in read mode
file_obj = open(inputPath+inputFile, 'r')
file_obj2 = openpyxl.load_workbook(inputPath+inputFile)
file_obj.buffer
file_obj2.sheetnames

# check if file is closed
fileClosed = checkFileClosed(file_obj)
print(f"file_obj.closed return >>>>> {fileClosed}")

# check if file is closed
fileClosed2 = checkFileClosedOpnxlpy(file_obj2)
print(f"OPNXLPY file_obj.closed return >>>>> {fileClosed2}")

print(f'Now closing {inputPath+inputFile} file.' )
# close the file
file_obj.close()

# again check if file is closed
fileClosed =  checkFileClosed(file_obj)
print(f"file_obj.closed return >>>>> {fileClosed}")

# close the file
file_obj2.close()
# check if file is closed
fileClosed2 = checkFileClosedOpnxlpy(file_obj2)
print(f"OPNXLPY file_obj.closed return >>>>> {fileClosed2}")