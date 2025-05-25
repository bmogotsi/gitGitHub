#! python3
# A Shebang indicates which Bash or Python interpreter to use to interpret an executable file. 
# https://app.datacamp.com/learn/tutorials/python-excel-tutorial
import sys, webbrowser, re, time , os
from pathlib import Path
# add a day to a date field
from datetime import datetime, timedelta, date
import openpyxl


inputPath = "C:/Users/ben/OneDrive/Documents/Gaba_Docs/Whisperer/FBRef/2025 Files/AllDomesticLeagues/"
inputFile = "Fbref_20250102"   + ".xlsx"

outputPath = "C:/Users/ben/OneDrive/Documents/Gaba_Docs/development/Python/Python_Ben/pgmOutput/"
outputFile = "WSB_20250102a" + ".xlsx"

iterate =[]
iterate.append('one')

for ite in iterate:

    try:
        # Load workbook
        wb = openpyxl.load_workbook(inputPath+inputFile)

        # Select active sheet
        # sheet = wb.active

        # Select specific sheet 
        ws = wb['Fbref_Teams']

        # count the number of rows and columns in this worksheet
        print('Total number of rows: '+str(ws.max_row)+'. And total number of columns: '+str(ws.max_column))

        # Access cell values / retrieve data from a specific cell 
        cell_value = ws['B2'].value

        # retrieve values from MULTIPLE COLUMNS (cells)
        values = [ws.cell(row=1,column=i).value for i in range(1,ws.max_column+1)]
        print(values)

        # retrieve MULTIPLE ROWS
        data=[ws.cell(row=i,column=3).value for i in range(2,12)]
        print(data) 

        # reading data from a range of cells (from column 1 to 6) AND (row 1 to 11)

        my_list = list()

        for value in ws.iter_rows(
            min_row=1, max_row=11, min_col=1, max_col=6, 
            values_only=True):
            my_list.append(value)
                    # left (<) indent columns {:<3}{:<10}{:<25}{:<30}{:<8}{:<3}
        for ele1,ele2,ele3,ele4,ele5,ele6 in my_list:
            (print ("{:<3}{:<10}{:<25}{:<30}{:>7}{:>3}".format(ele1,ele2,ele3,ele4,ele5,ele6)))

        wb.close(inputPath+inputFile)
        break

        wRow = ws.max_row
        wCol = ws.max_column
        wCell = ws.cell(4,5).value
        values = [ws.cell(row=1,column=i).value for i in range(1,ws.max_column+1)] # all columns for row 1
        print(values)
        valuesR = [ws.cell(row=i,column=3).value for i in range(2,ws.max_row+1)] #all rows for column 3 ## All teams Array/List
        print(valuesR)
        mapTeam = [ws.cell(row=i,column=26).value for i in range(2,ws.max_row+1)] #all rows for column 3 ## All teams Array/List column Z mapped WSB name
        print(mapTeam)
    except Exception as e:
        print("could not open input file..: " + inputPath+inputFile+'\n' + inputPath + '\n' + inputFile + '\n' +str(e))

exit