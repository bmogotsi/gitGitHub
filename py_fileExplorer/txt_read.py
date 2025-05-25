#! python3
# A Shebang indicates which Bash or Python interpreter to use to interpret an executable file.
# convert TXT file to a XLSX worksheet 
# https://stackoverflow.com/questions/43007979/how-to-import-txt-file-into-python
# https://stackoverflow.com/questions/35028683/python3-unicodedecodeerror-with-readlines-method


# import csv
import openpyxl
# from openpyxl.cell import get_column_letter
from openpyxl.utils.cell import get_column_letter
import pathlib
import unicodedata


inputPath = "C:/Users/ben/OneDrive/Documents/Gaba_Docs/Whisperer/FBRef/2025 Files/AllDomesticLeagues/"
inputFile = "FBref_20250115_Text"   + ".txt"
# inputFile = "FBref_20250115_Csv_"   + ".csv"

outputPath = "C:/Users/ben/OneDrive/Documents/Gaba_Docs/development/Python/Python_Ben/pgmOutput/"
outputFile = "WSB_20250102x" + ".xlsx"

dest_filename = outputPath+outputFile
wb = openpyxl.load_workbook(dest_filename)

sheetNew="WSB_Teams"
# Delete
wb.remove(wb[sheetNew])
wb.create_sheet(sheetNew) # create an empty sheet 
ws=wb[sheetNew]

while 2 > 1:
    try:
        #ws = wb.worksheets[0]
        #ws.title = "WSB_Teams"
        # open in Python 3 allows you to provide the known encoding of an input, replacing the default (ASCII in your case) with any other recognized encoding.
        with open(inputPath+inputFile, encoding='utf-8', mode='r') as f:
            for row_index, line in enumerate(f.readlines()):
                l = line.strip().split(',')
                for column_index, cellv in enumerate(l):
                    if column_index == 2:
                        cellvAscii = str(unicodedata.normalize('NFKD', cellv).encode('ascii', 'ignore')).strip("b").strip("'")
                    else:
                        cellvAscii =cellv
                    column_letter = get_column_letter((column_index + 1))
                    #ws.cell('%s%s'%(column_letter, (row_index + 1))).value = cellv
                    ws.cell((row_index + 1),(column_index + 1)).value = cellvAscii 

                l = line.strip().split(',')
                array1 = l[0]
                array2 = l[1]

    except Exception as e:
        print("could not open input file..: " + inputPath+inputFile+'\n' + inputPath + '\n' + inputFile + '\n' +str(e))

    wb.save(filename = dest_filename)

    file_extension = pathlib.Path(outputPath+outputFile).suffix
    len1=len(str('prjHTMParse_Selinium_Chrome_WSB_output'))
    if inputFile[0:38]=='prjHTMParse_Selinium_Chrome_WSB_output':
        print(str(38))
    if inputFile[0:37]=='prjHTMParse_Selinium_Chrome_WSB_output':
        print(str(37))
    if inputFile[0:len1]=='prjHTMParse_Selinium_Chrome_WSB_output':
        print(str(len1)+ 'str(len1)')
    if inputFile[0:len1-1]=='prjHTMParse_Selinium_Chrome_WSB_output':
        print(str(len1-1) + 'str(len1-1)')
    break

exit