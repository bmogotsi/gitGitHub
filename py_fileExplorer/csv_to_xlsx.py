#! python3
# A Shebang indicates which Bash or Python interpreter to use to interpret an executable file.
# convert CSV file to a XLSX worksheet 
# https://stackoverflow.com/questions/35028683/python3-unicodedecodeerror-with-readlines-method

import csv
import openpyxl
# from openpyxl.cell import get_column_letter
from openpyxl.utils.cell import get_column_letter
import pathlib
import unicodedata


inputPath = "C:/Users/ben/OneDrive/Documents/Gaba_Docs/Whisperer/FBRef/2025 Files/AllDomesticLeagues/"
inputFile = "FBref_Csv_"   + ".csv"
# inputFile = "FBref_20250115_Text"   + ".txt"

outputPath = "C:/Users/ben/OneDrive/Documents/Gaba_Docs/Whisperer/FBRef/2025 Files/AllDomesticLeagues/"
outputFile = "FBref_Teams" + ".xlsx"

# open in Python 3 allows you to provide the known encoding of an input, replacing the default ('cp1252' in my case) with any other recognized encoding.
f = open(inputPath+inputFile, encoding='utf-8', mode='r')

csv.register_dialect('commas', delimiter=',')

reader = csv.reader(f, dialect='commas')

dest_filename = outputPath+outputFile
wb = openpyxl.load_workbook(dest_filename)

sheetNew="FBref_Teams" # case sensitive
# Delete
wb.remove(wb[sheetNew])
wb.create_sheet(sheetNew) # create an empty sheet 
ws=wb[sheetNew]

#ws = wb.worksheets[0]
#ws.title = "WSB_Teams"

for row_index, row in enumerate(reader):
    for column_index, cellv in enumerate(row):
        if column_index == 2:
            cellvAscii = str(unicodedata.normalize('NFKD', cellv).encode('ascii', 'ignore')).strip("b").strip("'")
        else:
            cellvAscii =cellv

        # Python program to check if a string has at least one letter and one numberb
        iSdigits= all(c.isdigit() for c in cellvAscii)
        iString= any(c.isalpha() for c in cellvAscii)

        # correct formating by changing field type
        # because the heading is alpha
        if iSdigits == True and cellvAscii != '': # INT
            cellvAscii=int(cellvAscii)
        else:
            try:
                cellvAscii = float(cellvAscii)
            except ValueError:
                cellvAscii = str(cellvAscii)

        column_letter = get_column_letter((column_index + 1))
        #ws.cell('%s%s'%(column_letter, (row_index + 1))).value = cellv
        ws.cell((row_index + 1),(column_index + 1)).value = cellvAscii 

wb.save(filename = dest_filename)

exit