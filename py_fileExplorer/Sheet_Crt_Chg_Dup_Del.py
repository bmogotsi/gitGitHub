#! python3
# A Shebang indicates which Bash or Python interpreter to use to interpret an executable file. 
# https://app.datacamp.com/learn/tutorials/python-excel-tutorial
import sys, webbrowser, re, time , os
from pathlib import Path
# add a day to a date field
from datetime import datetime, timedelta, date
import openpyxl
import random
num = random.random()


inputPath = "C:/Users/ben/OneDrive/Documents/Gaba_Docs/Whisperer/FBRef/2025 Files/AllDomesticLeagues/"
inputFile = "Fbref_20250102"   + ".xlsx"

outputPath = "C:/Users/ben/OneDrive/Documents/Gaba_Docs/development/Python/Python_Ben/pgmOutput/"
outputFile = "WSB_20250102a" + ".xlsx"

iterate =[]
iterate.append('one')

for ite in iterate:

    try:
        # Load workbook
        wb = openpyxl.load_workbook(inputPath+inputFile)

        # Select active sheet
        ws = wb.active

        # Select specific sheet 
        # ws = wb['Fbref_Teams']

        # count the number of rows and columns in this worksheet
        print('Total number of rows: '+str(ws.max_row)+'. And total number of columns: '+str(ws.max_column))

        # Access cell values / retrieve data from a specific cell 
        cell_value = ws['B2'].value

        # duplicate worksheet
        wb.copy_worksheet(wb['Sheet1'])
        (print ("{:>10}{:>55}".format("After >>>" , str(wb.sheetnames))))
        # break

        # Deleting  a New
        print("Active >>>     " + ws.title)
        (print ("{:>10}{:>55}".format("Before >>>" ,  str(wb.sheetnames))))

        answer = [i for i in wb.sheetnames if 'Copy' in i] #return if part of a string is contain

        print("<<<<< To be Deleted >>>>>>")
        for sheet0 in answer:
            print(sheet0)

            # Delete
            wb.remove(wb[sheet0])

        break
        # Print Sheet Name(S)
        # print("Before >>>     " + ws.title)
        (print ("{:>10}{:>55}".format("Before >>>" ,  str(wb.sheetnames))))
        
        # Creat Sheet Name
        sheetNew = 'Empty Sheet' + str(random.random())
        wb.create_sheet(sheetNew) # create an empty sheet  
        (print ("{:>10}{:>55}".format("After >>>" , str(wb.sheetnames))))

        
        # break

        # Creating a New Worksheet

        # Print Sheet Name(S)
        # print("Before >>>     " + ws.title)
        (print ("{:>10}{:>55}".format("Before >>>" ,  str(wb.sheetnames))))
        
        # Creat Sheet Name
        sheetNew = 'Empty Sheet' + str(random.random())
        wb.create_sheet(sheetNew) # create an empty sheet  
        (print ("{:>10}{:>55}".format("After >>>" , str(wb.sheetnames))))


        # Creating a New Worksheet

        # Print Active Sheet Name
        # print("Before >>>     " + ws.title)
        (print ("{:>10}{:>15}".format("Before >>>" ,  ws.title)))

        # Change Sheet Name
        if ws.title == 'Fbref_Teams':
            ws.title ='Fbref_Changed'
        else:
            ws.title ='Fbref_Teams'

        (print ("{:>10}{:>15}".format("After >>>" , ws.title)))




    except Exception as e:
        print("could not open input file..: " + inputPath+inputFile+'\n' + inputPath + '\n' + inputFile + '\n' +str(e))

wb.save(inputPath+inputFile)
wb.close()
exit