#! Python3
#
import openpyxl

inputPath = "C:/Users/ben/OneDrive/Documents/Gaba_Docs/Whisperer/FBRef/2025 Files/AllDomesticLeagues/"
inputFile = "FBref_Teams"   + ".xlsx"

# ----------------------------Function Begin--------sheet_DeleteRows____------------------------
def delete_ExcelRows(sheetName, startRow, endRow):
    sheetName.delete_rows(startRow, endRow)
# ----------------------------Function END--------sheet_DeleteRows____------------------------


file_obj = openpyxl.load_workbook(inputPath+inputFile)
# Select active sheet (optional)
sheet_matched = file_obj['SearchDelete']
searchString = str(sheet_matched['A2'].value).strip()

print(f'Maximum Rows B4 delete....:  {sheet_matched.max_row}')

beginAtRow = 3
delete_ExcelRows(sheet_matched, beginAtRow, sheet_matched.max_row-(beginAtRow-1))

file_obj.save(inputPath+inputFile)