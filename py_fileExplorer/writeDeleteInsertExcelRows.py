#! python3
# A Shebang indicates which Bash or Python interpreter to use to interpret an executable file. 
import sys, webbrowser, re, time , os
from pathlib import Path
# add a day to a date field
from datetime import datetime, timedelta, date
import openpyxl


inputPath = "C:/Users/ben/OneDrive/Documents/Gaba_Docs/Whisperer/FBRef/2025 Files/AllDomesticLeagues/"
inputFile = "Fbref_20250102"   + ".xlsx"

outputPath = "C:/Users/ben/OneDrive/Documents/Gaba_Docs/development/Python/Python_Ben/pgmOutput/"
outputFile = "WSB_20250102a" + ".xlsx"

iterate =[]
iterate.append('one')

for ite in iterate:

    try:
        # Load workbook
        wb = openpyxl.load_workbook(inputPath+inputFile)

        # Select active sheet
        # sheet = wb.active

        # Select specific sheet 
        ws = wb['Fbref_Teams']

        # Writing to a Cell
        ws['K1'] = 'Points per Match' # Pts/MP 

        # Or Alternatively row,col
        ws.cell(row=1, column=11, value = 'Points per Match')

        ws.cell(row=1, column=11, value = 'Pts/MP')

        # Creating a New Column
        row_position = 2
        col_position = 7

        total_MatchPlayed = ((ws.cell(row=row_position, column=col_position).value)+
                    (ws.cell(row=row_position, column=col_position+1).value)+
                    (ws.cell(row=row_position, column=col_position+2).value))

                
        ws.cell(row=1, column=28, value = 'Total Match Played')
        ws.cell(row=2,column=28).value=total_MatchPlayed
        
        # create a for loop to sum the Total Match Played in every row
        row_position = 1

        for i in range(1, ws.max_row):

            row_position += 1
            MP_Win  = ws.cell(row=row_position, column=7).value
            MP_Draw = ws.cell(row=row_position, column=8).value
            MP_Lose = ws.cell(row=row_position, column=9).value

            total_MP = (MP_Win + MP_Draw + MP_Lose)
            ws.cell(row=row_position, column=28).value = total_MP

        # Appending New Rows
        # To append a new row to the workbook, simply create a tuple with the values you’d like to include and write it to the sheet:
        insRow = (ws.max_row+1,'2024-2025','La LouviÃ¨re','be Challenger Pro League','2.20',21,13,6,2,33,'2.20','2nd',1.44,65,27,14,13,'41.7',5,'33.3',23,2,2,0,'La LouviÃ¨re','La Louviere', '????',21)
        # ws.append(insRow)

        # Deleting Rows
        # To delete the new row we just created, you can run the following line of code:
        # ws.delete_rows(ws.max_row, 1) # row number, number of rows to delete

        # SAVE workbook
        wb.save(inputPath+inputFile)
        wb.close()
        break


    except Exception as e:
        print("could not open input file..: " + inputPath+inputFile+'\n' + inputPath + '\n' + inputFile + '\n' +str(e))

exit