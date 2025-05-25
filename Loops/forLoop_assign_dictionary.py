#! python3
# A Shebang indicates which Bash or Python interpreter to use to interpret an executable file. 

# Technically, every function returns exactly one value; that value, however, can be a tuple, 
#           a list, or some other type that contains multiple values.
# That said, you can return something that uses something other than just the order of values to distinguish them. 
#           You can return a dict:

import openpyxl
import csv
from openpyxl.utils.cell import get_column_letter
import collections

inputPath = "C:/Users/ben/OneDrive/Documents/Gaba_Docs/Whisperer/FBRef/2025 Files/AllDomesticLeagues/"
inputFile = "Fbref_Teams"   + ".xlsx"

# ----------------------------Function BEGIN____FBref Sheet Values------------------------
iteration=True
while iteration==True:
    iteration=False
    try:
        def testFunction(sheet_Obj):
            # Access cell values
            cell_value = sheet_obj['B2'].value
            ref_Row = sheet_obj.max_row
            ref_Col = sheet_obj.max_column
            wCell = sheet_obj.cell(4,5).value
            ref_Headings = [sheet_obj.cell(row=1,column=i).value for i in range(1,sheet_obj.max_column+1)] # all columns for row 1 (Headings)
            # print(FBref_Headings)
            ref_Col_C = [sheet_obj.cell(row=i,column=3).value for i in range(2,sheet_obj.max_row+1)]  # all rows for column 3 ## All teams Array/List column C FBref name
            # print(FBref_Col_C)
            ref_Col_Z = [sheet_obj.cell(row=i,column=26).value for i in range(2,sheet_obj.max_row+1)] # all rows for column 26 ## All teams Array/List column Z FBref name
            # print(FBref_Col_Z)

            return dict(FBref_Headings=ref_Headings, FBref_Col_C=ref_Col_C, FBref_Col_Z=ref_Col_Z, FBref_Row=ref_Row, FBref_Col=ref_Col)

            # ----------------------------Function END____FBref Sheet Values------------------------

        # Opening the "Fbref_Teams" file in read mode
        file_obj = openpyxl.load_workbook(inputPath+inputFile)
        sheet_obj = file_obj['FBref_Teams']

        x = testFunction(sheet_obj)
        print(x['FBref_Headings'])
        print(x['FBref_Col_C'])
        print(x['FBref_Col_Z'])
        print(x['FBref_Row'])
        print(x['FBref_Row'])

        # -------------------------------------named tuple-------------------------------------------------------------------
        # or you can define a named tuple:
        ReturnType = collections.namedtuple('ReturnMultiple', 'FBref_Headings FBref_Col_C FBref_Col_Z FBref_Row FBref_Col')

        # ----------------------------Function BEGIN____FBref Sheet Values--------named tuple----------------
        def testFunction2(sheet_obj):
                # Access cell values
                cell_value = sheet_obj['B2'].value
                ref_Row = sheet_obj.max_row
                ref_Col = sheet_obj.max_column
                wCell = sheet_obj.cell(4,5).value
                ref_Headings = [sheet_obj.cell(row=1,column=i).value for i in range(1,sheet_obj.max_column+1)] # all columns for row 1 (Headings)
                # print(FBref_Headings)
                ref_Col_C = [sheet_obj.cell(row=i,column=3).value for i in range(2,sheet_obj.max_row+1)]  # all rows for column 3 ## All teams Array/List column C FBref name
                # print(FBref_Col_C)
                ref_Col_Z = [sheet_obj.cell(row=i,column=26).value for i in range(2,sheet_obj.max_row+1)] # all rows for column 26 ## All teams Array/List column Z FBref name
                # print(FBref_Col_Z)

                return ReturnType(FBref_Headings=ref_Headings, FBref_Col_C=ref_Col_C, FBref_Col_Z=ref_Col_Z, FBref_Row=ref_Row, FBref_Col=ref_Col)

                # ----------------------------Function END____FBref Sheet Values------------------------

        y = testFunction2(sheet_obj)
        print(y.FBref_Col_C)  # or x[0], if you still want to use the index
        print(y.FBref_Col_Z) 
        print(y.FBref_Row) 
        print(y.FBref_Row) 
        print(y.FBref_Headings) 

        # -------------------------------------End Of ____________ named tuple-------------------------------------------------------------------

    except Exception as e:
        print("Something Went Wrong "+ str(e))


    exit